import tkinter as tk
from PIL import Image, ImageTk
import openpyxl 
from datetime import date
import os

def wyczysc():
    # Czyści prawy panel
    for widget in prawy_panel.winfo_children():
        widget.destroy()

    # Odświeżenie tła 
    tlo_prawe = tk.Label(prawy_panel, image=tlo_prawy_panel)
    tlo_prawe.place(x=0, y=0, relwidth=1, relheight=1)
    tlo_lewe = tk.Label(lewy_panel, image=tlo_lewy_panel)
    tlo_lewe.place(x=0, y=0, relwidth=1, relheight=1)

#---------------------------------------------------------------------------------LISTA PRODUKTÓW
def akcja_1():
    wyczysc()

    # Pole wyszukiwania
    szukaj_frame = tk.Frame(lewy_panel, bg="#261d1c")
    szukaj_frame.place(x=20, y=0, width=560, height=30)

    szukaj_var = tk.StringVar()
    szukaj_entry = tk.Entry(szukaj_frame, textvariable=szukaj_var, font=("Arial", 12))
    szukaj_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=5, pady=2)

    # Funkcja filtrująca listę po nazwie zaczynającej się od wpisanej frazy
    def filtruj_liste(event=None):
        lista.delete(0, tk.END)
        fraza = szukaj_var.get().lower()
        for row in dane_z_excela:
            if row[1].lower().startswith(fraza):
                tekst = f"ID: {row[0]}  |  Produkt: {row[1]}  |  Pojemność: {row[2]}"
                lista.insert(tk.END, tekst)

    szukaj_entry.bind('<Return>', filtruj_liste)
    
    filtruj_btn = tk.Button(szukaj_frame, text="Szukaj", command=filtruj_liste, font=("Arial", 10), bg="#b3685b", fg="#261d1c", activebackground="#453735")
    filtruj_btn.pack(side=tk.RIGHT, padx=5)

    # Utworzenie listy
    lista_frame = tk.Frame(lewy_panel, bg="#261d1c")
    lista_frame.place(x=20, y=30, width=560, height=450)

    scrollbar = tk.Scrollbar(lista_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    lista = tk.Listbox(lista_frame, yscrollcommand=scrollbar.set, font=("Arial", 12), bg="#b3685b", fg="#261d1c")
    lista.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=lista.yview)

    # Wczytanie danych z Excela
    global dane_z_excela
    dane_z_excela = []

    try:
        wb = openpyxl.load_workbook("produkty.xlsx")
        arkusz = wb.active

        for row in arkusz.iter_rows(min_row=2, max_col=3, values_only=True):
            if row[0] and row[1] and row[2]:
                dane_z_excela.append(row)
                tekst = f"ID: {row[0]}  |  Produkt: {row[1]}  |  Pojemność: {row[2]}"
                lista.insert(tk.END, tekst)

    except FileNotFoundError:
        lista.insert(tk.END, "Nie znaleziono pliku 'produkty.xlsx'")
    except Exception as e:
        lista.insert(tk.END, f"Błąd: {str(e)}")

    #Dodaj do bazy
    dodaj_tekst = tk.Label(prawy_panel, text="---------Dodaj do bazy---------", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    dodaj_tekst.place(x=20, y=30)

    id_tekst = tk.Label(prawy_panel, text="ID:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_tekst.place(x=20, y=60)
    pole_tekstowe_id = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_id.place(x=20, y=90)

    produkt_tekst = tk.Label(prawy_panel, text="Produkt:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    produkt_tekst.place(x=20, y=120)
    pole_tekstowe_produkt = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_produkt.place(x=20, y=150)

    pojemnosc_tekst = tk.Label(prawy_panel, text="Pojemnosc:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    pojemnosc_tekst.place(x=20, y=180)
    pole_tekstowe_pojemnosc = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_pojemnosc.place(x=20, y=210)

    waga_tekst = tk.Label(prawy_panel, text="Waga butelki:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    waga_tekst.place(x=20, y=240)
    pole_tekstowe_waga = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_waga.place(x=20, y=270)


    usun_tekst = tk.Label(prawy_panel, text="---------Usuń z bazy---------", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    usun_tekst.place(x=20, y=360)
    id_tekst = tk.Label(prawy_panel, text="ID:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_tekst.place(x=20, y=390)
    pole_tekstowe_id_2 = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_id_2.place(x=20, y=420)

    def zapisz_do_excela():
        id_val = pole_tekstowe_id.get()
        produkt = pole_tekstowe_produkt.get()
        pojemnosc_val = pole_tekstowe_pojemnosc.get()
        waga_val = pole_tekstowe_waga.get()


        if not id_val or not waga_val or not produkt or not pojemnosc_val:
            print('Nie wszystkie pola zostały uzupełnione!')
            return
        
        plik = 'produkty.xlsx'

        if os.path.exists(plik):
            wb = openpyxl.load_workbook(plik)
            arkusz = wb.active
        
        data = date.today()

        arkusz.append([id_val, produkt, pojemnosc_val, waga_val, data])
        wb.save(plik)
        zapisano = tk.Label(prawy_panel, text='Pomyślnie zapisano', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
        zapisano.place(x=170, y=300)
    
    przycisk_zapisz = tk.Button(prawy_panel, text="Zatwierdź dodanie", bg="#b3685b", fg="#261d1c", activebackground="#453735", font=("Arial", 12), command=zapisz_do_excela)
    przycisk_zapisz.place(x=20, y=300)

    def usun_z_bazy():
        id_val = pole_tekstowe_id_2.get()
        plik = 'produkty.xlsx'
        wb = openpyxl.load_workbook(plik)
        arkusz = wb.active

        for row in range(2, arkusz.max_row + 1):
            id_z_pliku = str(arkusz.cell(row=row, column=1).value)
            if id_z_pliku == str(id_val):
                arkusz.delete_rows(row)
                wb.save(plik)
                usunieto = tk.Label(prawy_panel, text='Pomyślnie usunięto', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
                usunieto.place(x=180, y=450)
                return
        nie_znaleziono = tk.Label(prawy_panel, text='Nie znaleziono produktu', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
        nie_znaleziono.place(x=180, y=450)

    przycisk_usun = tk.Button(prawy_panel, text="Zatwierdź usunięcie", bg="#b3685b", fg="#261d1c", activebackground="#453735", font=("Arial", 12), command=usun_z_bazy)
    przycisk_usun.place(x=20, y=450)

#---------------------------------------------------------------------------------DODANIE INWENTARYZACJI
def akcja_2():
    wyczysc()
    #Id
    id_tekst = tk.Label(prawy_panel, text="ID produktu:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_tekst.place(x=20, y=30)
    pole_tekstowe_id = tk.Entry(prawy_panel, width=30, bg="#b3685b") 
    pole_tekstowe_id.place(x=20, y=60)


    #Waga
    waga_tekst = tk.Label(prawy_panel, text="Waga (butelka + płyn):", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    waga_tekst.place(x=20, y=100)
    pole_tekstowe_waga = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_waga.place(x=20, y=130)

    #Ilość pełnych butelek
    butelki_tekst = tk.Label(prawy_panel, text="Pełne butelki:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    butelki_tekst.place(x=20, y=170)
    pole_tekstowe_butelki = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_butelki.place(x=20, y=200)

    def zapisz_do_excela():
        id_val = pole_tekstowe_id.get()
        waga_val = pole_tekstowe_waga.get()
        butelki_val = pole_tekstowe_butelki.get()
        data = date.today()

        if not id_val or not waga_val or not butelki_val:
            print('Nie wszystkie pola zostały uzupełnione!')
            return
        
        plik = 'inwentaryzacja.xlsx'

        if os.path.exists(plik):
            wb = openpyxl.load_workbook(plik)
            arkusz = wb.active
        else:
            wb = openpyxl.Workbook()
            arkusz = wb.active
            arkusz.append(['ID', 'Data', 'Waga produktu', 'Pełne butelki','Waga płynu'])
        

        plik2 = 'produkty.xlsx'#ŁĄCZENIE BAZ PO ID

        if os.path.exists(plik2):
            wb_laczenie = openpyxl.load_workbook(plik2)
            arkusz_laczenie = wb_laczenie.active


        szukane_id = pole_tekstowe_id.get()

        for row in arkusz_laczenie.iter_rows(min_row=2, values_only=True):
            id_wiersza, nazwa = row[0], row[1]
            if str(id_wiersza) == szukane_id:
                print(f"Nazwa dla ID={szukane_id}: {nazwa}")

                break
        else:
            print(f"Nie znaleziono ID={szukane_id}")

        arkusz.append([id_val, data, waga_val, butelki_val, nazwa])
        wb.save(plik)
        zapisano = tk.Label(prawy_panel, text='Pomyślnie zapisano', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
        zapisano.place(x=20, y=280)

    przycisk_zapisz = tk.Button(prawy_panel, text="Zatwierdź dodanie", bg="#b3685b", fg="#261d1c", activebackground="#453735", font=("Arial", 12), command=zapisz_do_excela)
    przycisk_zapisz.place(x=20, y=240)



#---------------------------------------------------------------------------------EDYCJA INWENTARYZACJI
def akcja_3():
    wyczysc()
    #Id
    id_tekst = tk.Label(prawy_panel, text="ID produktu podlegającego zmianie:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_tekst.place(x=20, y=30)
    pole_tekstowe_id = tk.Entry(prawy_panel, width=30, bg="#b3685b") 
    pole_tekstowe_id.place(x=20, y=60)

    #Nowa poprawna waga
    waga_tekst = tk.Label(prawy_panel, text="Nowa, poprawna waga:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    waga_tekst.place(x=20, y=100)
    pole_tekstowe_waga = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_waga.place(x=20, y=130)

    #Ilość pełnych butelek
    butelki_tekst = tk.Label(prawy_panel, text="Poprawna ilość pełnych butelek:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    butelki_tekst.place(x=20, y=170)
    pole_tekstowe_butelki = tk.Entry(prawy_panel, width=30, bg="#b3685b")
    pole_tekstowe_butelki.place(x=20, y=200)

    def edycja():
        id_val = pole_tekstowe_id.get()
        waga = pole_tekstowe_waga.get()
        butelki_val = pole_tekstowe_butelki.get()

        plik = 'inwentaryzacja.xlsx'
        wb = openpyxl.load_workbook(plik)
        arkusz = wb.active

        if not id_val or not waga or not butelki_val:
            print('Nie wszystkie pola zostały uzupełnione!')
            return
        
        for row in range(2, arkusz.max_row + 1):
            id_z_pliku = str(arkusz.cell(row=row, column=1).value)
            if id_z_pliku == str(id_val):
                arkusz.cell(row=row, column=3).value = waga
                arkusz.cell(row=row, column=4).value = butelki_val
                wb.save(plik)

                zmieniono = tk.Label(prawy_panel, text='Pomyślnie zmieniono', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
                zmieniono.place(x=20, y=280)
                return
        nie_znaleziono = tk.Label(prawy_panel, text='Nie znaleziono produktu', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
        nie_znaleziono.place(x=20, y=280)

    zamiana = tk.Button(prawy_panel, text="Zatwierdź zmianę", bg="#b3685b", fg="#261d1c", activebackground="#453735", font=("Arial", 12), command=edycja)
    zamiana.place(x=20, y=250)

        
#---------------------------------------------------------------------------------USUWANIE INWENTARYZACJI 
def akcja_4():
    wyczysc()

    #Id
    id_usuwanego = tk.Label(prawy_panel, text="ID usuwanego produktu:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_usuwanego.place(x=20, y=30)
    wejscie_id_usuwanego = tk.Entry(prawy_panel, width=30, bg="#b3685b") 
    wejscie_id_usuwanego.place(x=20, y=60)


    def usun_z_inwentaryzacji():
        id_usuwane = wejscie_id_usuwanego.get()
        plik = 'inwentaryzacja.xlsx'
        wb = openpyxl.load_workbook(plik)
        arkusz = wb.active

        for row in range(2, arkusz.max_row + 1):
            id_z_pliku = str(arkusz.cell(row=row, column=1).value)
            if id_z_pliku == str(id_usuwane):
                arkusz.delete_rows(row)
                wb.save(plik)
                usunieto = tk.Label(prawy_panel, text='Pomyślnie usunięto', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
                usunieto.place(x=20, y=150)
                return
        nie_znaleziono = tk.Label(prawy_panel, text='Nie znaleziono produktu', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
        nie_znaleziono.place(x=20, y=150)
        
    przycisk_usun_z_bazy = tk.Button(prawy_panel, text="Zatwierdź usunięcie", bg="#b3685b", fg="#261d1c", activebackground="#453735", font=("Arial", 12), command=usun_z_inwentaryzacji)
    przycisk_usun_z_bazy.place(x=20, y=100)


#---------------------------------------------------------------------------------STATYSTYKI
def akcja_5():
    wyczysc()

    #Wyszukiwanie po ID
    id_tekst = tk.Label(prawy_panel, text="ID produktu:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    id_tekst.place(x=20, y=30)
    pole_tekstowe = tk.Entry(prawy_panel, width=30, bg="#b3685b") 
    pole_tekstowe.place(x=20, y=60)

    #Przewidywana waga
    przewidywana_waga_tekst = tk.Label(prawy_panel, text='Przewidywana waga produktu:', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    przewidywana_waga_tekst.place(x=20, y=100)
    pole_tekstowe5 = tk.Entry(prawy_panel, width=30, bg='#b3685b')
    pole_tekstowe5.place(x=20, y=130)

    #Przewidywana ilość butelek
    przewidywana_il_butl_tekst = tk.Label(prawy_panel, text='Przewidywana ilość butelek:', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    przewidywana_il_butl_tekst.place(x=20, y=170)
    pole_tekstowe6 = tk.Entry(prawy_panel, width=30, bg='#b3685b')
    pole_tekstowe6.place(x=20, y=200)

    #Faktyczna waga
    faktyczna_waga_tekst = tk.Label(prawy_panel, text="Faktyczna waga produktu:", bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    faktyczna_waga_tekst.place(x=20, y=240)

    #Faktyczna ilość butelek
    faktyczna_il_butl_tekst = tk.Label(prawy_panel, text='Faktyczna ilość butelek:', bg="#261d1c", fg="#b3685b", font=("Arial", 14))
    faktyczna_il_butl_tekst.place(x=20, y=310)
#---------------------------------------------------------------------------------OKNO APLIKACJI
okno = tk.Tk()
okno.title("Program inwentaryzacji baru")#nazwa aplikacji
okno.geometry("1000x600")#wymiary okna - najlepiej nie zmieniac

#---------------------------------------------------------------------------------PANELE (WYMIARY, USTAWIENIE)
#menu
gorny_panel = tk.Frame(okno, width=1000, height=70, bg="#291612", bd=2, relief="raised", highlightthickness=3, highlightbackground="black")
gorny_panel.place(x=0, y=0)

#lewy panel (tylko poglądowy)
lewy_panel = tk.Frame(okno, width=600, height=530, highlightthickness=3, highlightbackground="black")
lewy_panel.place(x=0, y=70)

#prawy panel (interaktywny)
prawy_panel = tk.Frame(okno, width=400, height=530, highlightthickness=3, highlightbackground="black")
prawy_panel.place(x=600, y=70)


#---------------------------------------------------------------------------------PANELE (USTAWIENIE GRAFIK)
grafika_lewy_panel = Image.open("tlol2.jpeg")
grafika_lewy_panel = grafika_lewy_panel.resize((600, 600), Image.Resampling.LANCZOS)
tlo_lewy_panel = ImageTk.PhotoImage(grafika_lewy_panel)

grafika_prawy_panel = Image.open("tlop.jpg")
grafika_prawy_panel = grafika_prawy_panel.resize((400, 600), Image.Resampling.LANCZOS)
tlo_prawy_panel = ImageTk.PhotoImage(grafika_prawy_panel)

grafika_gorny_panel = grafika_prawy_panel.resize((1000, 70), Image.Resampling.LANCZOS)
tlo_gorny_panel = ImageTk.PhotoImage(grafika_gorny_panel)

tlo_lewe = tk.Label(lewy_panel, image=tlo_lewy_panel)
tlo_lewe.place(x=0, y=0, relwidth=1, relheight=1)

tlo_prawe = tk.Label(prawy_panel, image=tlo_prawy_panel)
tlo_prawe.place(x=0, y=0, relwidth=1, relheight=1)

gorny_panel = tk.Label(gorny_panel, image=tlo_gorny_panel)
gorny_panel.place(x=0, y=0, relwidth=1, relheight=1)

#---------------------------------------------------------------------------------PRZYCISKI, MOŻNA DODAĆ WIĘCEJ - NIE WYSYPIE SIE
btn1 = tk.Button(gorny_panel, text="Lista produktów", command=akcja_1, bg="#261d1c", fg="#b3685b", activebackground="#453735")
btn1.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

btn2 = tk.Button(gorny_panel, text="Nowa inwentaryzacja", command=akcja_2, bg="#261d1c", fg="#b3685b", activebackground="#453735")
btn2.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

btn3 = tk.Button(gorny_panel, text="Edytuj inwentaryzacje", command=akcja_3, bg="#261d1c", fg="#b3685b", activebackground="#453735")
btn3.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

btn4 = tk.Button(gorny_panel, text="Usuń inwentaryzacje", command=akcja_4, bg="#261d1c", fg="#b3685b", activebackground="#453735")
btn4.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

btn5 = tk.Button(gorny_panel, text="Wyświetl statystyki", command=akcja_5, bg="#261d1c", fg="#b3685b", activebackground="#453735")
btn5.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

okno.mainloop()

